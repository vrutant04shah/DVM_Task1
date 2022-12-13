import openpyxl

path = "C:/Users/vruta/OneDrive/Desktop/sutt.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet = wb_obj.active

loops = sheet.max_row

dictionary = []
i = 1
for i in range(1, loops):
    dict = {'name': f"{sheet.cell(i + 1, 1).value}", 'author': f"{sheet.cell(i + 1, 2).value}",
            'isbn': f'{sheet.cell(i + 1, 3).value}', 'genre': f"{sheet.cell(i + 1, 4).value}",
            'status': f"{sheet.cell(i + 1, 5).value}"}
    dictionary.append(dict)


class Book:
    def __init__(self, name):
        self.name = name

    def Issue_book(self):  # For issuing a book from the list of books
        for ite in dictionary:
            if self.name == ite['name']:
                if ite['status'] == 'Available':
                    print(f"{name_book} has been issued by you.")
                    ite['status'] = 'Issued'
                else:
                    print(f"Sorry, {name_book} is not Available right now.")
                break

    def Return_book(self):  # For returning a book which is already issued or reserved
        for val in dictionary:
            if self.name == val['name']:
                if val['status'] == 'Issued' or val['status'] == 'Reserved':
                    print(f"{name_book} has been returned by you.")
                    val['status'] = 'Available'
                elif val['status'] == 'Available':
                    print(f"The book - {name_book} has not been issued by you.")
                break

    def Reserve_book(self):  # For reserving a book from the list of books
        for ele in dictionary:
            if self.name == ele['name']:
                if ele['status'] == 'Available':
                    print(f"The book {name_book} has been reserved by you.")
                    ele['status'] = 'Reserved'
                else:
                    print("Sorry, the book is already in use.")


class Shelf:
    def __init__(self, name):
        self.name = name

    def populate_book(self):  # For populating the list of books from external excel file.
        print(loops2)
        for t in range(1, loops2):
            nul = {"name": f"{sheet1.cell(t + 1, 1).value}",
                   "author": f"{sheet1.cell(t + 1, 2).value}",
                   "isbn": f"{sheet1.cell(t + 1, 3).value}",
                   "genre": f"{sheet1.cell(t + 1, 4).value}",
                   "status": "Available"}
            dictionary.append(nul)
            sheet.cell(loops + t, 1).value = f"{sheet1.cell(t + 1, 1).value}"
            sheet.cell(loops + t, 2).value = f"{sheet1.cell(t + 1, 2).value}"
            sheet.cell(loops + t, 3).value = f"{sheet1.cell(t + 1, 3).value}"
            sheet.cell(loops + t, 4).value = f"{sheet1.cell(t + 1, 4).value}"
            sheet.cell(loops + t, 5).value = f"Available"
            wb_obj.save("C:/Users/vruta/OneDrive/Desktop/sutt.xlsx")

    def Add_book(self):  # For adding a book to the list of books
        author = input("Enter the name of the author: ")
        isbn = input("Enter the isbn number of the book: ")
        genre = input("Enter the genre of the book: ")
        status = 'Available'
        dict_add = {"name": f"{self.name}", "author": f"{author}", "isbn": f"{isbn}",
                    "genre": f"{genre}", "status": f"{status}"}
        dictionary.append(dict_add)
        print("The book has been successfully added to the list of books.")
        sheet.cell(loops + 1, 1).value = f"{self.name}"
        sheet.cell(loops + 1, 2).value = f"{author}"
        sheet.cell(loops + 1, 3).value = f"{isbn}"
        sheet.cell(loops + 1, 4).value = f"{genre}"
        sheet.cell(loops + 1, 5).value = f"Available"
        wb_obj.save("C:/Users/vruta/OneDrive/Desktop/sutt.xlsx")

    def Remove_book(self):  # For removing the book from the list of books
        for book_rem in dictionary:
            if self.name == book_rem["name"]:
                bb = book_rem
                t = dictionary.index(bb)
                dictionary.pop(t)
                print("The book has been successfully removed from the list of books.")
                break

    def Edit_book(self):  # For editing the details of the book
        if detail == 'name':
            print(f"The current value of the name of the book is {self.name}.\n")
            self.name = input("What would you like the book's name to be? ")
            print("\nThe book's name has been successfully changed.\n")
        elif detail == 'isbn':
            for ele in dictionary:
                if self.name == ele['name']:
                    print(f"The current isbn number of {self.name} is {ele['isbn']}.\n")
                    ele['isbn'] = input("What would you like to change isbn to? ")
                    print(f"\nThe isbn of {self.name} has been successfully changed.\n")
        elif detail == 'author':
            for ele in dictionary:
                if self.name == ele['name']:
                    print(f'The current author of the book is {ele["author"]}')
                    ele["author"] = input("What would you like to change the author to? ")
                    print(f"The name of the author for the book,{self.name} has been successfully changed.")
        elif detail == 'genre':
            for ele in dictionary:
                if self.name == ele['name']:
                    print(f'The current author of the book is {ele["genre"]}')
                    ele["genre"] = input("What would you like to change the genre to? ")
                    print(f"The genre of the book,{self.name} has been successfully changed.")
                    break
        else:
            print("Enter proper detail to make changes!!")


class Info:
    def __init__(self):
        print(*dictionary, sep="\n")


def get_book_count():
    print(f"The total number of Books in the shelf is {sheet.max_row - 1}.")


# Starting with User type, I have assigned different options to the user.
# After defining User, I have used OOPS to conduct different requirements of the User.

while True:
    user = input("Which kind of user are you ? (Basic/Librarian) : ")
    if user.lower() == 'basic' or user.lower() == 'librarian':
        break
    else:
        print("Please enter correct user information!")

print("Welcome to the Library!!")

while True:
    if user.lower() == 'basic':
        user_work = input("For issuing book, enter 1.\n"
                          "For returning book, enter 2.\n"
                          "For reserving book, enter 3.\n"
                          "For getting information, enter 'info'.\n"
                          "OPTION: ")
        if user_work.lower() == 'info':
            Info()
        elif user_work == '1':
            name_book = input("Please enter the name of the book : ")
            Book1 = Book(name_book)
            Book.Issue_book(Book1)
        elif user_work == '2':
            name_book = input("Please enter the name of the book : ")
            Book2 = Book(name_book)
            Book.Return_book(Book2)
        elif user_work == '3':
            name_book = input("Please enter the name of the book : ")
            Book3 = Book(name_book)
            Book.Reserve_book(Book3)
        elif user_work.lower() == 'exit':
            break
        else:
            print("Please enter a proper option!")

    elif user.lower() == 'librarian':
        librarian_work = input("What work would you like to do:\n"
                               "For adding a book, enter 1.\n"
                               "For removing a book, enter 2.\n"
                               "For editing a book, enter 3.\n"
                               "To get total number of books in Library, enter 4.\n"
                               "To get information on books, enter 5.\n"
                               "For adding book from other file(excel), enter 6.\n"
                               "To Exit, enter exit.\n"
                               "OPTION: ")
        if librarian_work == '1':
            new_book = input("Enter the name of the book : ")
            Book4 = Shelf(new_book)
            Shelf.Add_book(Book4)
        elif librarian_work == '2':
            rem_book = input("Enter the name of the book to be removed : ")
            Book5 = Shelf(rem_book)
            Shelf.Remove_book(Book5)
        elif librarian_work == '3':
            while True:
                edit_book = input("Enter the name of the book which has to be edited : ")
                for item in dictionary:
                    if edit_book == item['name']:
                        detail = input("\nEnter which detail you would like to change:(name/author/isbn/genre) ")
                        Book6 = Shelf(edit_book)
                        Shelf.Edit_book(Book6)
                        break
        elif librarian_work == '4':
            get_book_count()
        elif librarian_work == '5':
            Info()
        elif librarian_work == '6':
            file = input("Enter the path of the file: ")
            wb_obj1 = openpyxl.load_workbook(file)
            sheet1 = wb_obj1.active
            loops2 = sheet1.max_row
            cell = sheet1[1][2]
            Shelf.populate_book(cell)
        elif librarian_work.lower() == 'exit':
            break
